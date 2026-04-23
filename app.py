import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from io import BytesIO
from PIL import Image
import matplotlib.pyplot as plt
from reportlab.platypus import SimpleDocTemplate, Image as RLImage
import tempfile

# ==============================
# CONFIGURATION DE LA PAGE
# ==============================
st.set_page_config(
    page_title="BOM Comparator", 
    page_icon="📊", 
    layout="wide",
    initial_sidebar_state="collapsed"
)

# ==============================
# STYLES CSS PERSONNALISÉS
# ==============================
st.markdown("""
    <style>
    .main-header {
        font-size: 2.5rem;
        font-weight: bold;
        text-align: center;
        color: #1E3A8A;
        margin-bottom: 1rem;
    }
    .sub-header {
        font-size: 1.2rem;
        text-align: center;
        color: #6B7280;
        margin-bottom: 2rem;
    }
    
    /* Style pour les statuts encadrés */
    .status-badge {
        display: inline-block;
        padding: 4px 12px;
        border-radius: 6px;
        font-weight: bold;
        text-align: center;
        border: 2px solid;
        margin: 2px 0;
    }
    .status-conform {
        background-color: #10B98120;
        color: #065F46;
        border-color: #10B981;
    }
    .status-missing {
        background-color: #EF444420;
        color: #991B1B;
        border-color: #EF4444;
    }
    .status-packing {
        background-color: #3B82F620;
        color: #1E3A8A;
        border-color: #3B82F6;
    }
    .status-qty {
        background-color: #F59E0B20;
        color: #92400E;
        border-color: #F59E0B;
    }
    .status-refchange {
        background-color: #8B5CF620;
        color: #5B21B6;
        border-color: #8B5CF6;
    }
    
    /* Style pour les lignes sélectionnables */
    .stDataFrame {
        cursor: pointer;
    }
    .selected-row {
        background-color: #EFF6FF !important;
        border-left: 4px solid #3B82F6 !important;
    }
    </style>
""", unsafe_allow_html=True)

# ==============================
# LOGO ET TITRE
# ==============================
col1, col2, col3 = st.columns([1, 2, 1])
with col2:
    try:
        logo = Image.open("logo.jfif")
        st.image(logo, width=200)
    except:
        pass

st.markdown('<div class="main-header">📊 BOM Comparator</div>', unsafe_allow_html=True)
st.markdown('<div class="sub-header">Outil de comparaison BOM vs Packing List</div>', unsafe_allow_html=True)

# ==============================
# SIDEBAR POUR LES INPUTS
# ==============================
with st.sidebar:
    st.markdown("## 📂 Upload des fichiers")
    bom_file = st.file_uploader("📄 BOM", type=["xlsx", "xls"], key="bom")
    packing_file = st.file_uploader("📦 Packing List", type=["xlsx", "xls"], key="packing")
    
    st.markdown("---")
    st.markdown("## 📝 Informations")
    model_input = st.text_input("📺 Modèle", placeholder="Ex: iPhone 13")
    lot_input = st.text_input("🔢 Quantité Lot", placeholder="Ex: 1000")
    
    st.markdown("---")
    run = st.button("🚀 Lancer la comparaison", use_container_width=True, type="primary")

# ==============================
# FONCTIONS UTILITAIRES
# ==============================
def format_status_with_border(status):
    """Retourne le statut avec un encadré HTML"""
    status_map = {
        "✅ Conform": '<span class="status-badge status-conform">✅ Conform</span>',
        "❌ Missing item": '<span class="status-badge status-missing">❌ Missing item</span>',
        "📦 Packing only": '<span class="status-badge status-packing">📦 Packing only</span>',
        "⚠ Qty missing": '<span class="status-badge status-qty">⚠ Qty missing</span>',
        "🔄 Reference Change": '<span class="status-badge status-refchange">🔄 Reference Change</span>'
    }
    return status_map.get(status, status)

def show_kpis(df):
    """Affiche les KPIs avec des cartes stylisées"""
    total = len(df)
    conform = (df["Remark"] == "✅ Conform").sum()
    missing = (df["Remark"] == "❌ Missing item").sum()
    packing_only = (df["Remark"] == "📦 Packing only").sum()
    qty_missing = (df["Remark"] == "⚠ Qty missing").sum()
    ref_change = (df["Remark"] == "🔄 Reference Change").sum()
    
    st.markdown("### 📈 Indicateurs de Performance")
    
    col1, col2, col3, col4, col5, col6 = st.columns(6)
    col1.metric("📊 Total", total)
    col2.metric("✅ Conformes", conform, delta=f"{conform/total*100:.1f}%" if total > 0 else "0%")
    col3.metric("❌ Manquants", missing, delta=f"-{missing/total*100:.1f}%" if total > 0 else "0%", delta_color="inverse")
    col4.metric("📦 Packing Only", packing_only, delta=f"{packing_only/total*100:.1f}%" if total > 0 else "0%")
    col5.metric("⚠️ Quantités", qty_missing, delta=f"{qty_missing/total*100:.1f}%" if total > 0 else "0%", delta_color="inverse")
    col6.metric("🔄 Ref Change", ref_change, delta=f"{ref_change/total*100:.1f}%" if total > 0 else "0%")
    
    return total, conform, missing, packing_only, qty_missing, ref_change

def generate_pie_chart(df):
    """Génère un graphique circulaire amélioré"""
    conform = (df["Remark"] == "✅ Conform").sum()
    missing = (df["Remark"] == "❌ Missing item").sum()
    packing_only = (df["Remark"] == "📦 Packing only").sum()
    qty_missing = (df["Remark"] == "⚠ Qty missing").sum()
    ref_change = (df["Remark"] == "🔄 Reference Change").sum()
    
    labels = ["✅ Conform", "❌ Missing", "📦 Packing Only", "⚠️ Qty Missing", "🔄 Ref Change"]
    values = [conform, missing, packing_only, qty_missing, ref_change]
    colors = ['#10B981', '#EF4444', '#3B82F6', '#F59E0B', '#8B5CF6']
    explode = (0.05, 0.05, 0.05, 0.05, 0.05)
    
    fig, ax = plt.subplots(figsize=(8, 6))
    wedges, texts, autotexts = ax.pie(
        values,
        labels=labels,
        autopct='%1.1f%%',
        startangle=90,
        colors=colors,
        explode=explode,
        shadow=True,
        textprops={'fontsize': 10, 'fontweight': 'bold'}
    )
    
    for autotext in autotexts:
        autotext.set_color('white')
        autotext.set_fontsize(11)
        autotext.set_fontweight('bold')
    
    ax.set_title("Distribution des Articles", fontsize=14, fontweight='bold', pad=20)
    plt.tight_layout()
    
    return fig

def color_remark_column(val):
    """Colorie la colonne Remark avec bordure"""
    colors = {
        "✅ Conform": "background-color: #10B98120; color: #065F46; font-weight: bold; border: 2px solid #10B981; border-radius: 6px; padding: 4px 8px;",
        "⚠ Qty missing": "background-color: #F59E0B20; color: #92400E; font-weight: bold; border: 2px solid #F59E0B; border-radius: 6px; padding: 4px 8px;",
        "❌ Missing item": "background-color: #EF444420; color: #991B1B; font-weight: bold; border: 2px solid #EF4444; border-radius: 6px; padding: 4px 8px;",
        "📦 Packing only": "background-color: #3B82F620; color: #1E3A8A; font-weight: bold; border: 2px solid #3B82F6; border-radius: 6px; padding: 4px 8px;",
        "🔄 Reference Change": "background-color: #8B5CF620; color: #5B21B6; font-weight: bold; border: 2px solid #8B5CF6; border-radius: 6px; padding: 4px 8px;"
    }
    return colors.get(val, "")

def detect_remark(row):
    """Détection automatique du statut"""
    if row["_merge"] == "left_only":
        return "❌ Missing item"
    elif row["_merge"] == "right_only":
        return "📦 Packing only"
    elif row["packing_qty"] >= row["Qty (MP+SAV)"]:
        return "✅ Conform"
    else:
        return "⚠ Qty missing"

def find_reference_changes(df):
    """Trouve les paires potentielles de changements de référence"""
    missing_items = df[df["Remark"] == "❌ Missing item"]
    packing_only = df[df["Remark"] == "📦 Packing only"]
    
    potential_pairs = []
    
    for _, missing in missing_items.iterrows():
        desc_missing = str(missing.get("Description", "")).lower()
        
        for _, packing in packing_only.iterrows():
            desc_packing = str(packing.get("Description", "")).lower()
            
            # Si les descriptions sont similaires, suggestion de paire
            if desc_missing and desc_packing and (
                desc_missing in desc_packing or 
                desc_packing in desc_missing or
                any(word in desc_packing for word in desc_missing.split()[:2])
            ):
                potential_pairs.append({
                    "old_ref": missing["PN"],
                    "new_ref": packing["PN"],
                    "description": desc_missing
                })
                break
    
    return potential_pairs

def export_excel(df):
    """Exporte vers Excel avec mise en forme"""
    output = BytesIO()
    
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Résultats")
    
    output.seek(0)
    wb = load_workbook(output)
    ws = wb.active
    
    color_map = {
        "✅ Conform": "C6EFCE",
        "⚠ Qty missing": "FFEB9C",
        "❌ Missing item": "FFC7CE",
        "📦 Packing only": "BDD7EE",
        "🔄 Reference Change": "E6D0FF"
    }
    
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        remark_cell = row[8]
        if remark_cell.value in color_map:
            color = color_map[remark_cell.value]
            for cell in row:
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    final = BytesIO()
    wb.save(final)
    final.seek(0)
    return final

# ==============================
# LOGIQUE PRINCIPALE
# ==============================
if run:
    if not bom_file or not packing_file:
        st.error("⚠️ Veuillez uploader les deux fichiers")
        st.stop()
    
    if not model_input:
        st.error("⚠️ Veuillez entrer un modèle")
        st.stop()
    
    if not lot_input.isdigit():
        st.error("⚠️ La quantité lot doit être un nombre")
        st.stop()
    
    lot = int(lot_input)
    
    with st.spinner("📖 Lecture des fichiers..."):
        bom = pd.read_excel(bom_file)
        packing = pd.read_excel(packing_file)
        
        bom.columns = bom.columns.str.strip()
        packing.columns = packing.columns.str.strip()
        
        packing["Model"] = packing["Model"].astype(str).str.strip()
        packing["Model"] = packing["Model"].replace("", None).ffill()
    
    packing_model = packing[packing["Model"] == model_input]
    
    if packing_model.empty:
        st.error(f"⚠️ Modèle '{model_input}' non trouvé dans le fichier Packing")
        st.stop()
    
    with st.spinner("🔄 Traitement des données..."):
        bom_g = bom.groupby(["PN", "Description"])["bom_qty"].sum().reset_index()
        packing_g = packing_model.groupby(["PN", "Description"])["packing_qty"].sum().reset_index()
        
        df = pd.merge(
            bom_g,
            packing_g,
            on="PN",
            how="outer",
            suffixes=("_BOM", "_Packing"),
            indicator=True
        )
        
        df["bom_qty"] = pd.to_numeric(df["bom_qty"], errors="coerce").fillna(0)
        df["packing_qty"] = pd.to_numeric(df["packing_qty"], errors="coerce").fillna(0)
        df["Description_BOM"] = df["Description_BOM"].fillna(df["Description_Packing"])
        
        df["MP"] = df["bom_qty"] * lot
        df["SAV"] = df["MP"] * 0.02
        df["Qty (MP+SAV)"] = df["MP"] + df["SAV"]
        df["Balance"] = df["packing_qty"] - df["Qty (MP+SAV)"]
        df["Remark"] = df.apply(detect_remark, axis=1)
        
        result = df[[
            "PN", "Description_BOM", "bom_qty", "packing_qty", 
            "MP", "SAV", "Qty (MP+SAV)", "Balance", "Remark"
        ]].rename(columns={
            "Description_BOM": "Description",
            "bom_qty": "Qty BOM",
            "packing_qty": "Packing list qty"
        })
    
    st.session_state["result"] = result
    st.session_state["data_ready"] = True
    st.session_state["selected_rows"] = []
    st.session_state["ref_changes"] = {}

# ==============================
# GESTION DES CHANGEMENTS DE RÉFÉRENCE
# ==============================
if "data_ready" in st.session_state and st.session_state["data_ready"]:
    result = st.session_state["result"].copy()
    
    # Appliquer les changements de référence déjà enregistrés
    if "ref_changes" in st.session_state:
        for old_ref, new_ref in st.session_state["ref_changes"].items():
            result.loc[result["PN"] == old_ref, "Remark"] = "🔄 Reference Change"
            result.loc[result["PN"] == new_ref, "Remark"] = "🔄 Reference Change"
    
    st.balloons()
    st.success("✅ Comparaison terminée avec succès !")
    
    # KPIs
    total, conform, missing, packing_only, qty_missing, ref_change = show_kpis(result)
    
    # Détection automatique des paires potentielles
    potential_pairs = find_reference_changes(result)
    
    if potential_pairs:
        st.info(f"💡 {len(potential_pairs)} paire(s) potentielle(s) de changement de référence détectée(s)")
        
        with st.expander("🔍 Suggestions de changements de référence"):
            for pair in potential_pairs:
                st.write(f"**Ancienne référence:** {pair['old_ref']} → **Nouvelle référence:** {pair['new_ref']}")
                st.write(f"Description: {pair['description']}")
                st.markdown("---")
    
    # Interface de sélection manuelle
    st.markdown("### 🎯 Gestion manuelle des changements de référence")
    st.markdown("Sélectionnez un article '❌ Missing item' et un article '📦 Packing only' pour les marquer comme changement de référence")
    
    # Afficher les colonnes disponibles pour la sélection
    missing_items = result[result["Remark"] == "❌ Missing item"]
    packing_items = result[result["Remark"] == "📦 Packing only"]
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown("#### ❌ Articles manquants (ancienne référence)")
        if not missing_items.empty:
            selected_missing = st.selectbox(
                "Sélectionner l'ancienne référence",
                options=missing_items["PN"].tolist(),
                format_func=lambda x: f"{x} - {missing_items[missing_items['PN']==x]['Description'].values[0][:50]}",
                key="missing_select"
            )
        else:
            st.info("Aucun article manquant détecté")
            selected_missing = None
    
    with col2:
        st.markdown("#### 📦 Articles Packing only (nouvelle référence)")
        if not packing_items.empty:
            selected_packing = st.selectbox(
                "Sélectionner la nouvelle référence",
                options=packing_items["PN"].tolist(),
                format_func=lambda x: f"{x} - {packing_items[packing_items['PN']==x]['Description'].values[0][:50]}",
                key="packing_select"
            )
        else:
            st.info("Aucun article Packing only détecté")
            selected_packing = None
    
    # Bouton pour valider le changement de référence
    if selected_missing and selected_packing:
        col_button1, col_button2, col_button3 = st.columns([1, 2, 1])
        with col_button2:
            if st.button("🔄 Marquer comme changement de référence", use_container_width=True, type="primary"):
                if "ref_changes" not in st.session_state:
                    st.session_state["ref_changes"] = {}
                
                st.session_state["ref_changes"][selected_missing] = selected_packing
                
                st.success(f"✅ Changement de référence enregistré : {selected_missing} → {selected_packing}")
                st.rerun()
    
    # Afficher les changements déjà effectués
    if "ref_changes" in st.session_state and st.session_state["ref_changes"]:
        st.markdown("#### 📝 Changements de référence effectués")
        
        changes_df = pd.DataFrame([
            {"Ancienne référence": old, "Nouvelle référence": new}
            for old, new in st.session_state["ref_changes"].items()
        ])
        st.dataframe(changes_df, use_container_width=True)
        
        if st.button("🗑️ Réinitialiser tous les changements", use_container_width=True):
            st.session_state["ref_changes"] = {}
            st.rerun()
    
    st.markdown("---")
    
    # Créer deux colonnes pour le graphique et le tableau
    col_left, col_right = st.columns([1, 2])
    
    with col_left:
        st.markdown("### 📊 Distribution")
        fig = generate_pie_chart(result)
        st.pyplot(fig, use_container_width=True)
        
        # Bouton PDF
        img_buffer = BytesIO()
        fig.savefig(img_buffer, format="png", dpi=150, bbox_inches='tight')
        img_buffer.seek(0)
        
        pdf_buffer = BytesIO()
        doc = SimpleDocTemplate(pdf_buffer)
        
        with tempfile.NamedTemporaryFile(delete=False, suffix=".png") as tmp:
            tmp.write(img_buffer.getvalue())
            tmp_path = tmp.name
        
        elements = [RLImage(tmp_path, width=350, height=350)]
        doc.build(elements)
        pdf_buffer.seek(0)
        
        st.download_button(
            "📄 Télécharger graphique (PDF)",
            data=pdf_buffer,
            file_name="KPI_Chart.pdf",
            mime="application/pdf",
            use_container_width=True
        )
    
    with col_right:
        st.markdown("### 📋 Détail des articles")
        
        # Fonction pour appliquer le style avec les bordures
        def style_dataframe(df):
            styled = df.style
            
            # Appliquer le style à la colonne Remark
            for idx, row in df.iterrows():
                status = row["Remark"]
                color = {
                    "✅ Conform": "background-color: #10B98120; color: #065F46;",
                    "⚠ Qty missing": "background-color: #F59E0B20; color: #92400E;",
                    "❌ Missing item": "background-color: #EF444420; color: #991B1B;",
                    "📦 Packing only": "background-color: #3B82F620; color: #1E3A8A;",
                    "🔄 Reference Change": "background-color: #8B5CF620; color: #5B21B6;"
                }.get(status, "")
                
                if color:
                    styled = styled.set_properties(**{
                        'background-color': color.split(';')[0].split(':')[1],
                        'color': color.split(';')[1].split(':')[1],
                        'font-weight': 'bold',
                        'border': '2px solid',
                        'border-radius': '6px',
                        'padding': '4px 8px'
                    }, subset=['Remark'])
            
            # Formater les nombres
            numeric_cols = ['Qty BOM', 'Packing list qty', 'MP', 'SAV', 'Qty (MP+SAV)', 'Balance']
            styled = styled.format({col: '{:.0f}' for col in numeric_cols if col in df.columns})
            
            return styled
        
        styled_df = style_dataframe(result)
        st.dataframe(styled_df, use_container_width=True, height=400)
    
    # Section export
    st.markdown("---")
    col_export1, col_export2, col_export3 = st.columns([1, 1, 1])
    
    with col_export2:
        excel_file = export_excel(result)
        st.download_button(
            "📥 Télécharger rapport Excel",
            data=excel_file,
            file_name=f"BOM_vs_Packing_{model_input}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            type="primary"
        )
